import requests
import time
from lxml import etree
from openpyxl import Workbook
from selenium import webdriver

wb = Workbook()  # 创建工作簿
ws = wb.active  # 激活工作表
driver = webdriver.Chrome()
# 头标
TITLE = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1',
         'S1', 'T1', 'U1', 'V1', 'W1', 'X1', 'Y1', 'Z1', 'AA1', 'AB1', 'AC1', 'AD1', 'AE1', 'AF1', 'AG1', 'AH1', 'AI1',
         'AJ1', 'AK1', 'AL1', 'AM1', 'AN1']
SUBJECTS = [101, 201, 204, 301, 302, 303, 306, 307, 311, 312, 397, 497, 398, 498, 199, 396]
# 待爬路由地址
DATAURL = []
excel_data = {}  # 存放在 excel 表中的数据
rowId = 2  # excel_data 的索引值


# 添加待爬路由地址
def set_url():
    DATAURL.append('http://yz.kaoyan365.cn/school/gzhu/zhuanye/')


# 将数据输出到excel表中
# def output_excel(str_data):
#     for index in range(len(PROVINCE)):
#         if str_data.find(PROVINCE[index]) > -1:
#             excel_data[PROVINCE[index]][TITLE[index]] += 1
#             ws.cell(excel_data[PROVINCE[index]][TITLE[index]], index + 1).value = str_data


# 获取页面中的文本数据
def get_webtext(site):
    # 获取页面中的文本数据
    url = site
    # 添加请求头部字段: 反反爬
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Referer': 'https://gz.zu.anjuke.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36',
    }
    cook = {
        'Cookie': 'lps=https%3A%2F%2Fgz.zu.anjuke.com%2F%3Fpi%3Dbaidu-cpchz-gz-cty1%26kwid%3D154696710660%26bd_vid%3D8678955026479273642%7Chttps%3A%2F%2Fwww.baidu.com%2Fbaidu.php%3Furl%3DKf0000K5cNxA6dzipXPx18obsXRMgwjd_2966kFwj2fJDYEC0xKrmYNmYzFwYm4FUDacjD0JiYRbi6JZV9XeoowQlEL9Gd7XGSB0O4O48aA9KVhFcl7-XPE3qAzkF1SOCwX3_V-iWGbrVFz2CN2hFPWUdMxMdxDGZX2EbI5VJfqEA6oYp4dcuXs9YgkrOxPwQnnAO4catZoTVobvUTS0VcrlSEDv.7D_NR2Ar5Od66dT5-s35FhQQ6eRC6e3L_g_3_AXZZjNs4qgjM9Cp2SEUsmh2qLuYeSMko_lXhki_nYQZubolpd0.U1Yk0ZDqzqZQqeBqExzdkQXOdSQfkTSB0ZKGm1Yk0Zfq8X5yLIxBVeOgYsKGUHYznWR0u1dEuZCk0ZNG5yF9pywd0ZKGujYY0APGujY3P0KVIjYknjD4g1DsnHIxn10knfKopHYs0ZFY5Hn3nsKBpHYkPH9xnW0Yg1ckPdtdnjn0UynqnH6vnWfzP1R4n7tkrHDkPHDvP10zg1Dsn6KkTA-b5H00TyPGujYs0ZFMIA7M5H00mycqn7ts0ANzu1Ys0ZKs5H00UMus5H08nj0snj0snj00Ugws5H00uAwETjYs0ZFJ5H00uANv5gKW0AuY5H00TA6qn0KET1Ys0AFL5HDs0A4Y5H00TLCq0A71gv-bm1dsTzdMXh93XfKGuAnqiD4K0ZKCIZbq0Zw9ThI-IjY1nNt1nHFxnHR0IZN15Hmvn1D4P1bdPHfkrjDzrHRkPjT0ThNkIjYkPWnzn16LnWb4PHbk0ZPGujdhryRznAmYnW0snj0zmyD40AP1UHYdn19AwHTdrH7afH6dwj0Y0A7W5HD0TA3qn0KkUgfqn0KkUgnqn0KbugwxmLK95H00XMfqn0KVmdqhThqV5HKxn7ts0Aw9UMNBuNqsUA78pyw15HKxn7tznH0dnWcz0ZK9I7qhUA7M5H00uAPGujYs0ANYpyfqQHD0mgPsmvnqn0KdTA-8mvnqn0KkUymqn0KhmLNY5H00pgPWUjYs0A7buhk9u1Yk0Akhm1Ys0AwWmvfq0Zwzmyw-5HR4njckP0KBuA-b5H-DnHmknHfYnDF7fWf3wj9Anbc3wWfLf10kP1fknRmz0AqW5HD0mMfqn6KEmgwL5H00ULfqnfKETMKY5HcWnanznanzc1fYn1nYnHDvnznkrHRWnHbdQW0snj0snankc1cWnanVc108PWmYrHnzc1D8njTLP10s0Z7xIWYsQWnsg108njKxna3sn7tsQWc3g108nWIxna34P7tsQW0sg1Dzr0KBTdqsThqbpyfqn0KzUv-hUA7M5Hc0mLmq0A-1gvPsmHYs0APs5H00ugPY5H00mLFW5HR3Pjfv%26us%3Dnewvui%26word%3D%26ck%3D3355.1.117.347.179.344.172.393%26shh%3Dwww.baidu.com%26wd%3D%26bc%3D110101; aQQ_ajkguid=5DA0FA6D-FEDC-0294-6F3A-87FCF457CF7B; id58=CrIHp2FMQOm4n1jbAwNbAg==; sessid=4923C4F3-8D78-3BFF-5435-3F75403F3423; twe=2; _ga=GA1.2.562024657.1632457419; _gid=GA1.2.945207058.1632457419; 58tj_uuid=2b32f4d6-755a-48d1-9fdc-b5bbf850e5ea; als=0; fzq_h=d8175edc03921be22a8e84306f87607d_1632460809604_4c24bedd0517406eb9b02cc941dba09c_2018863672; new_session=1; new_uv=3; utm_source=; spm=; init_refer=https%253A%252F%252Fguangzhou.anjuke.com%252Fcommunity%252Fp1%252F; obtain_by=1; ctid=12; cmctid=3; xxzl_cid=f23f8b1a331f4a8a9f1f920c6722998b; xzuid=cc891613-1181-4772-8075-b44b9797a485'

    }
    response = requests.get(url=url, headers=headers, cookies=cook)
    response.encoding = 'utf-8'
    return response.text


def get_info2(site_data):
    global rowId
    page_text = get_webtext(site_data)
    tree = etree.HTML(page_text)
    data = tree.xpath('//div[@class="jibenMsg publicMsg"]/table/tr[9]/td[2]/text()')[0]
    data_list = data.split('(')[1:]
    for i in range(len(data_list)):
        k = 0  # 当k等于-1时，不存金excel
        for s in SUBJECTS:
            if data_list[i].find(str(s)) > -1:
                k = -1
                break
        if k == -1:
            continue
        else:
            # 存进excel表中
            list = data_list[i].split(')')
            print(list)
            if list:
                ws.cell(rowId, 2).value = list[1]
                ws.cell(rowId, 3).value = list[0]
                rowId += 1


# 获取数据并存进 excel 中
def get_info(site_data):
    page_text = get_webtext(site_data)
    # 解析页面数据并创建etree对象
    tree = etree.HTML(page_text)
    div_list = tree.xpath('//*[@class="zhuanYeMenu"]/div[@class="menuItem"]')
    # 如果是有效数据，则添加进 excel 表中
    for menu_item in div_list:
        a_list = menu_item.xpath('./div[@class="menuLinks clearfix"]/a')
        for a_item in a_list:
            print(a_item.text[8:])
            ws.cell(rowId, 1).value = a_item.text[8:]
            print('+++++++++++')
            site = 'http:' + a_item.xpath('./@href')[0]
            driver.get(site)
            get_info2(site)


# 入口
if __name__ == '__main__':
    print("开始爬取数据，请稍等...")
    set_url()
    strAdd = './data.xlsx'  # 本地数据文件地址
    start_time = time.time()

    for url in DATAURL:
        get_info(url)
    # 输出到 excel 表
    wb.save(strAdd)

    end_time = time.time()
    print("数据爬取完毕，用时%.2f秒" % (end_time - start_time))
