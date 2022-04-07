from openpyxl import Workbook

wb = Workbook()  # 创建工作簿
ws = wb.active  # 激活工作表
# ws.append(['111','',333])
# ws.append(['','222',333])
# ws.append(['','',333])
# wb.save('./test.xlsx')

def filter(str, citys):
    # 获取"市"索引
    for index in len(citys):
        if str.find(citys[index]) > -1:
            arr = []
            for i in range(index):
                arr.append('')
            arr.append(citys[index])
            print(arr)
            ws.append(arr)

citys = ['AAA', 'BBB', 'CCC']
# ws['A1'] = 'AAA'
# ws['B1'] = 'BBB'
# ws['C1'] = 'CCC'
# filter('asdasdlijBBBasd', citys)
# wb.save('./test.xlsx')
