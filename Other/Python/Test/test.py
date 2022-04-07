from lxml import etree
from openpyxl import Workbook

title = ['A1', 'B1', 'C1']
wb = Workbook()  # 创建工作簿
ws = wb.active  # 激活工作表

for t in title:
    ws[t] = t
# ws.append(title)
ws.cell(1, 1).value = 'abc'
ws['A5'] = 'AAAA'

wb.save('./test.xlsx')
